#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv
import re
import ast
import math
from datetime import datetime
from pathlib import Path
from collections import defaultdict

import numpy as np
from openpyxl import load_workbook

# =======================
# Fixed I/O
# =======================
INPUT_XLSX = Path("/data/hyan671/yhproject/FakingRecipe/lenslang/outputs/data_analysis/BOTH/RQ1/result/result_raw.xlsx")
OUTPUT_DIR = Path("/data/hyan671/yhproject/FakingRecipe/lenslang/outputs/data_analysis/BOTH/RQ1/result/output")


def _mean_std(xs):
    xs = [float(x) for x in xs if x is not None and str(x) != ""]
    if not xs:
        return float("nan"), float("nan")
    return float(np.mean(xs)), float(np.std(xs, ddof=0))


def _fmt_mu_sd(mu, sd, digits=4):
    if math.isnan(mu) or math.isnan(sd):
        return ""
    return f"{mu:.{digits}f}±{sd:.{digits}f}"


def _extract_metrics_dict(cell_value):
    """
    cell_value can be:
      - "{'auc': ..., 'f1': ..., 'acc': ...}"
      - "\"{'auc': ...}\n\"" (with quotes/newlines)
    Return a python dict or None.
    """
    if cell_value is None:
        return None
    s = str(cell_value).strip()

    # strip surrounding quotes (Excel sometimes keeps them)
    s = s.strip().strip('"').strip("'").strip()

    # find the first {...} block
    m = re.search(r"\{.*\}", s, flags=re.S)
    if not m:
        return None

    d_str = m.group(0)
    # parse dict safely (only Python literals)
    d = ast.literal_eval(d_str)
    if not isinstance(d, dict):
        return None
    return d


def _pick_sheet(wb):
    """
    Prefer sheets that look like raw input; otherwise active.
    """
    preferred = ["Raw", "raw", "result_raw", "ResultRaw", "Sheet1"]
    for name in preferred:
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def read_xlsx_as_bucket(xlsx_path: Path):
    """
    Read rows like:
        FakeSV | M1 | "{'auc':..., 'f1':..., 'acc':...}"
              | M2 | "{...}"
              | M3 | "{...}"
        FakeTT | M1 | "{...}"
              | M2 | "{...}"
              | M3 | "{...}"

    Return: bucket[dataset][model][metric] -> list of floats
    """
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = _pick_sheet(wb)

    bucket = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    current_ds = None
    for row in ws.iter_rows(values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        c0 = row[0] if len(row) > 0 else None
        c1 = row[1] if len(row) > 1 else None
        c2 = row[2] if len(row) > 2 else None

        # dataset may appear only on the first line of a block
        if c0 is not None and str(c0).strip() != "":
            ds_candidate = str(c0).strip()
            if ds_candidate in ("FakeSV", "FakeTT"):
                current_ds = ds_candidate

        model = str(c1).strip() if c1 is not None else ""
        if not current_ds or model not in ("M1", "M2", "M3"):
            continue

        d = _extract_metrics_dict(c2)
        if not d:
            continue

        # keys are expected: acc / f1 / auc (case-insensitive)
        d_lc = {str(k).lower(): v for k, v in d.items()}
        for k in ("acc", "f1", "auc"):
            if k in d_lc:
                try:
                    bucket[current_ds][model][k].append(float(d_lc[k]))
                except Exception:
                    pass

    return bucket


def build_summary_rows(bucket):
    """
    Output rows:
    Dataset, Model, Acc(mean±std), F1(mean±std), AUC(mean±std),
    ΔAcc(M2−M1), ΔF1(M2−M1), ΔAUC(M2−M1), Sig.
    """
    headers = [
        "Dataset", "Model",
        "Acc (mean±std)", "F1 (mean±std)", "AUC (mean±std)",
        "ΔAcc (M2−M1)", "ΔF1 (M2−M1)", "ΔAUC (M2−M1)",
        "Sig."
    ]

    order = [("FakeSV", ["M1", "M2", "M3"]), ("FakeTT", ["M1", "M2", "M3"])]

    out = [headers]
    for ds, models in order:
        m1_acc_mu, _ = _mean_std(bucket[ds]["M1"]["acc"])
        m2_acc_mu, _ = _mean_std(bucket[ds]["M2"]["acc"])
        m1_f1_mu,  _ = _mean_std(bucket[ds]["M1"]["f1"])
        m2_f1_mu,  _ = _mean_std(bucket[ds]["M2"]["f1"])
        m1_auc_mu, _ = _mean_std(bucket[ds]["M1"]["auc"])
        m2_auc_mu, _ = _mean_std(bucket[ds]["M2"]["auc"])

        for md in models:
            acc_mu, acc_sd = _mean_std(bucket[ds][md]["acc"])
            f1_mu,  f1_sd  = _mean_std(bucket[ds][md]["f1"])
            auc_mu, auc_sd = _mean_std(bucket[ds][md]["auc"])

            row = [
                ds,
                md,
                _fmt_mu_sd(acc_mu, acc_sd),
                _fmt_mu_sd(f1_mu, f1_sd),
                _fmt_mu_sd(auc_mu, auc_sd),
                "—", "—", "—",
                "—",
            ]

            # only fill deltas on M2
            if md == "M2" and (not math.isnan(m1_acc_mu)) and (not math.isnan(m2_acc_mu)):
                row[5] = f"{(m2_acc_mu - m1_acc_mu):.4f}"
                row[6] = f"{(m2_f1_mu  - m1_f1_mu):.4f}"
                row[7] = f"{(m2_auc_mu - m1_auc_mu):.4f}"

            out.append(row)

    return out


def write_csv(rows, out_csv: Path):
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    # newline='' is recommended by Python csv docs
    with out_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerows(rows)


def main():
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(f"Input not found: {INPUT_XLSX}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_csv = OUTPUT_DIR / f"RQ1_result_{ts}.csv"

    bucket = read_xlsx_as_bucket(INPUT_XLSX)
    rows = build_summary_rows(bucket)
    write_csv(rows, out_csv)
    print(f"[OK] Saved -> {out_csv}")


if __name__ == "__main__":
    main()